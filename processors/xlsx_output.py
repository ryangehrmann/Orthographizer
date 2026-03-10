"""
processors/xlsx_output.py
Build the formatted output .xlsx file from a processed DataFrame.

Column order: index, sub_index, entry_ortho, entry, word_lao, word,
              P, R, C, M, V, F, T, ambiguous, [original extra columns]

word_lao col: plain Lao-script value (empty for non-Lao pipelines).
word column:  Excel formula =<P>&<R>&<C>&<M>&<V>&<F>&<T> (columns computed
              dynamically), shaded gray, not intended for direct editing.

Column widths (in Excel character units):
  6  → index, sub_index
  24 → entry_ortho, entry
  12 → word_lao, word, ambiguous
  3  → P, R, C, M, V, F, T
"""
import io
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from core.modifiers import syllables_to_word


# ── Fixed output columns (before extra cols) ────────────────────────────────
FIXED_COLS = ["index", "sub_index", "entry_ortho", "entry", "word_lao", "word",
              "P", "R", "C", "M", "V", "F", "T", "ambiguous"]

# Segment slot columns whose values the 'word' formula concatenates
_SEG_COLS = ["P", "R", "C", "M", "V", "F", "T"]

COLUMN_WIDTHS = {
    "index": 6,
    "sub_index": 6,
    "entry_ortho": 24,
    "entry": 24,
    "word_lao": 12,
    "word": 12,
    "P": 3, "R": 3, "C": 3, "M": 3, "V": 3, "F": 3, "T": 3,
    "ambiguous": 12,
}

GRAY_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")

_FONT_DEFAULT = Font(name="Times New Roman", size=10)
_FONT_LAO     = Font(name="Saysettha OT",    size=12)


def _word_formula(row: int, seg_letters: list) -> str:
    """Return the Excel formula for the 'word' cell at the given 1-based row."""
    return "=" + "&".join(f"{letter}{row}" for letter in seg_letters)


def build_output_rows(
    df: pd.DataFrame,
    pipeline_run_fn,
    flat_dict: dict,
    segments_dict: dict,
) -> tuple[pd.DataFrame, dict]:
    """
    Process each entry in df through the pipeline, explode by syllable/word,
    and return:
      - output_df: DataFrame with all required columns (word column = placeholder)
      - stats: {"n_entries": int, "n_rows": int, "n_ambiguous": int, "unknown_chars": DataFrame}

    The 'word' column is left as empty string here; Excel formulas are written
    by write_xlsx().
    """
    from core.conversion import find_unknown_chars

    rows = []
    extra_cols = [c for c in df.columns if c not in ("entry_ortho", "index")]

    for _, row in df.iterrows():
        idx = row.get("index", "")
        entry_ortho = row.get("entry_ortho", "")
        extra = {col: row.get(col, "") for col in extra_cols}

        ipa, syllables = pipeline_run_fn(entry_ortho, flat_dict, segments_dict)

        # Rebuild the entry string from cleaned syllable forms so it reflects
        # the modifier-processed IPA rather than the raw conversion output.
        clean_entry = syllables_to_word(syllables) if syllables else (ipa or "")

        if not syllables:
            rows.append({
                "index": idx, "sub_index": 0,
                "entry_ortho": entry_ortho, "entry": clean_entry,
                "word_lao": "", "word": "",
                "P": "", "R": "", "C": "", "M": "", "V": "", "F": "", "T": "",
                "ambiguous": False,
                **extra,
            })
            continue

        for sub_idx, syl in enumerate(syllables):
            rows.append({
                "index": idx,
                "sub_index": sub_idx,
                "entry_ortho": entry_ortho,
                "entry": clean_entry,
                "word_lao": syl.get("word_lao", ""),
                "word": "",  # filled by Excel formula
                "P": syl.get("P", ""),
                "R": syl.get("R", ""),
                "C": syl.get("C", ""),
                "M": syl.get("M", ""),
                "V": syl.get("V", ""),
                "F": syl.get("F", ""),
                "T": syl.get("T", ""),
                "ambiguous": syl.get("ambiguous", False),
                **extra,
            })

    output_df = pd.DataFrame(rows)

    # Column order
    final_extra = [c for c in output_df.columns if c not in FIXED_COLS]
    output_df = output_df[FIXED_COLS + final_extra]

    # Stats
    n_ambiguous = int(output_df["ambiguous"].sum()) if "ambiguous" in output_df.columns else 0

    # Unknown char check (on the IPA output)
    from core.conversion import find_unknown_chars as _fuc
    # We check unknown chars on entry_ortho using the flat_dict
    unknown_chars = pd.DataFrame(columns=["char", "unicode"])  # placeholder

    stats = {
        "n_entries": len(df),
        "n_rows": len(output_df),
        "n_ambiguous": n_ambiguous,
    }

    return output_df, stats


def write_xlsx(output_df: pd.DataFrame) -> bytes:
    """
    Write output_df to an in-memory xlsx file with full formatting.
    Returns raw bytes suitable for st.download_button.
    """
    extra_cols = [c for c in output_df.columns if c not in FIXED_COLS]
    all_cols = FIXED_COLS + extra_cols

    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        output_df[all_cols].to_excel(writer, index=False, sheet_name="Output")
        ws = writer.sheets["Output"]

        n_data_rows = len(output_df)

        # ── Word column formulas ──────────────────────────────────────────
        word_col_idx = all_cols.index("word") + 1  # 1-based
        word_col_letter = get_column_letter(word_col_idx)

        # Compute segment column letters dynamically (handles word_lao shift)
        seg_letters = [get_column_letter(all_cols.index(c) + 1) for c in _SEG_COLS]

        for excel_row in range(2, n_data_rows + 2):  # skip header row
            ws[f"{word_col_letter}{excel_row}"].value = _word_formula(excel_row, seg_letters)

        # ── Gray shading on word column (header + data) ──────────────────
        for excel_row in range(1, n_data_rows + 2):
            ws[f"{word_col_letter}{excel_row}"].fill = GRAY_FILL

        # ── Column widths ────────────────────────────────────────────────
        for col_idx, col_name in enumerate(all_cols, start=1):
            letter = get_column_letter(col_idx)
            width = COLUMN_WIDTHS.get(col_name, 14)  # default 14 for extra cols
            ws.column_dimensions[letter].width = width

        # ── Freeze header row and enable autofilter ──────────────────────
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # ── Fonts and alignment ───────────────────────────────────────────
        for col_idx, col_name in enumerate(all_cols, start=1):
            letter = get_column_letter(col_idx)
            for excel_row in range(1, n_data_rows + 2):
                cell = ws[f"{letter}{excel_row}"]
                # Lao font for word_lao data cells; Times New Roman everywhere else
                if col_name == "word_lao" and excel_row > 1:
                    cell.font = _FONT_LAO
                else:
                    cell.font = _FONT_DEFAULT
                if col_name in ("P", "R", "C", "M", "V", "F", "T"):
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(wrap_text=False)

    return buf.getvalue()
